from datetime import date
import xlwt
from django import template
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.template import loader
from django.urls import reverse
import json
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.models import User
# from mysqlx import View

from zayed_university_app.utils import render_to_pdf
from .models import Report
from zayed_university_app.models import Log
import pyodbc
from .models import DepartmentAdminUser, Department, UserType
from django.views.decorators.cache import cache_control
from django.contrib.auth import authenticate, login, logout

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test

import pandas as pd
import openpyxl


def admin_check(user):
    if user.is_staff or user.is_superuser:
        return True
    return False


def page_not_found(request):
    return render(request, 'home/page-404.html')


def get_custuser(cname, cadmin, cdepart, chk_true):
    global chk_list
    if chk_true:
        chk_list.clear()
        chk_list.append(cname)
        chk_list.append(cadmin)
        chk_list.append(cdepart)
    else:
        # print("in get custuser ", chk_list)
        return chk_list


def get_connection():
    # # for server conn = pymssql.connect(server='192.168.5.79', user='chatboat_sa', password='ch@tb0@t$@',
    # database='zu_chatbot_log_devel')

    conn = pyodbc.connect('Driver=SQL Server;'  ## for local
                          'Server=ZAHEER;'
                          'Database=zu_chatbot_log_dev;'
                          'Trusted_Connection=yes;')
    return conn


def deptwise_data(dept_name):
    global busy_users, rep_users
    connection = ''
    dept_log_record = ''
    new_users = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()
        cursor1 = connection.cursor()
        cursor2 = connection.cursor()
        cursor3 = connection.cursor()

        SQL_select_Query = "SELECT * FROM fn_deptwise_ans_data('" + dept_name + "');"
        cursor.execute(SQL_select_Query)
        log_record = cursor.fetchall()
        dept_log_record = [list(tp) for tp in log_record]

        rep_select_Query = "SELECT * FROM fn_deptwise_repeated_users('" + dept_name + "');"
        cursor1.execute(rep_select_Query)
        rep_users = cursor1.fetchall()
        rep_users = [list(tp) for tp in rep_users]

        busy_select_Query = "SELECT * FROM fn_deptwise_busy_period_cnt('" + dept_name + "');"
        cursor2.execute(busy_select_Query)
        busy_users = cursor2.fetchall()
        busy_users = [list(tp) for tp in busy_users]

        new_select_Query = "SELECT * FROM fn_deptwise_new_usr_cnt('" + dept_name + "');"
        cursor3.execute(new_select_Query)
        new_users = cursor3.fetchall()
        new_users = [list(tp) for tp in new_users]
        # print("new_users- ", new_users)

    except pyodbc.Error as error:
        print("Error while fetching data from SQL", error)

    finally:
        if connection:
            cursor.close()
            connection.close()
            print("SQL connection is closed")

    return dept_log_record, rep_users, busy_users, new_users


def dept_chart(dept_name):
    data, rep_data, busy_data, new_user_data_ = deptwise_data(dept_name)

    event_categories = list()
    wrong_ans_data = list()
    right_ans_data = list()
    no_ans_data = list()

    for dt in data:

        if dt[1].__str__() not in event_categories:
            event_categories.append(dt[1].__str__())  # for answer description

    edict = {}

    for i in event_categories:

        temp = [0, 0, 0]

        if i not in edict.keys():

            for dt in data:

                if i == dt[1].__str__():

                    if dt[0] == 'Right Answer': temp[0] = dt[2]

                    if dt[0] == 'Wrong Answer': temp[1] = dt[2]

                    if dt[0] == 'No Answer': temp[2] = dt[2]

        edict[i] = temp

    # print(edict)

    for key, value in edict.items():
        right_ans_data.append(value[0])

        wrong_ans_data.append(value[1])

        no_ans_data.append(value[2])

    wrong_answer = {

        'name': 'wrong_answer',

        'data': wrong_ans_data,

        'color': '#e53935'

    }

    right_answer = {

        'name': 'right_answer',

        'data': right_ans_data,

        'color': '#43a047'

    }

    no_answer = {

        'name': 'no_answer',

        'data': no_ans_data,

        'color': '#fb8c00'

    }

    bar_chart = {

        'chart': {'type': 'column',

                  },

        'title': {'text': 'Log Summary(Bar Chart)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    line_chart = {

        'chart': {'type': 'line'},

        'title': {'text': 'Log Summary(Line Charts)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif',

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    bar_dump = json.dumps(bar_chart)

    line_dump = json.dumps(line_chart)

    return bar_dump, line_dump, rep_data, busy_data, new_user_data_


def get_repeated_bot_users():
    busy_count = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT * from dbo.repeated_bot_users_day_view order by ev_count desc;'
        cursor.execute(SQL_select_Query)

        busy_count = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count


def get_repeated_bot_monthly_users():
    busy_count = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT * from dbo.repeated_bot_users_monthly_view order by ev_count desc;'
        cursor.execute(SQL_select_Query)

        busy_count = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)

    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count


def get_busy_period_count():
    busy_count = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT * from dbo.busy_period_count order by intent_counts desc;'
        cursor.execute(SQL_select_Query)

        busy_count = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count


def get_busy_period_count_monthly():
    busy_count_month = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT * from dbo.busy_period_count_monthly order by intent_counts desc;'
        cursor.execute(SQL_select_Query)

        busy_count_month = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)

    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count_month


def live_count():
    live = []
    live_month = []
    connection = ''
    try:
        connection = get_connection()
        cursor = connection.cursor()

        cursor1 = connection.cursor()

        SQL_select_Query = 'Select Top 1 * from dbo.livechat_daywise order by day desc;'
        cursor.execute(SQL_select_Query)

        live = cursor.fetchall()
        SQL_select_Query1 = 'Select top 1 * from dbo.livechat_monthwise order by month desc;'
        cursor1.execute(SQL_select_Query1)

        live_month = cursor1.fetchall()

    except pyodbc.Error as error:
        print(error)

    finally:
        if connection:
            cursor.close()
            connection.close()

    return live, live_month


def reset_count():
    reset_count = []

    reset_month = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        cursor1 = connection.cursor()

        SQL_select_Query = 'Select Top 1 * from dbo.reset_daywise order by day desc;'
        cursor.execute(SQL_select_Query)

        reset_count = cursor.fetchall()

        # print("reset_count", reset_count)

        SQL_select_Query1 = 'Select top 1 * from dbo.reset_monthwise order by month desc;'
        cursor1.execute(SQL_select_Query1)

        reset_month = cursor1.fetchall()

        # print("reset_month", reset_month)

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

            print("SQL connection is closed")

    return reset_count, reset_month


def get_total_users_cnt():
    total_eng_users_cnt = 0

    total_new_users_cnt = 0
    connection = ''
    """ query data from the log_record table """

    try:

        connection = get_connection()

        cursor2 = connection.cursor()

        cursor1 = connection.cursor()

        # postgreSQL_select_Query = 'SELECT * from chatbot.total_cnt_view_mview;'

        eng_select_Query = 'SELECT * from dbo.engaged_users_view1;'

        cursor2.execute(eng_select_Query)

        new_select_Query = 'SELECT * from dbo.new_users_view1;'

        cursor1.execute(new_select_Query)

        # log_record = cursor.fetchone()

        total_eng_users_cnt = cursor2.fetchall()

        total_new_users_cnt = cursor1.fetchall()

        # print("total_eng_users_cnt---> ", total_eng_users_cnt)
        # #
        # print("total_new_users_cnt---> ", total_new_users_cnt)

        # total_new_users_cnt = [list(tp) for tp in desc_total_cnt]



    except pyodbc.Error as error:

        print("Error while fetching total_desc_cnt from SQL", error)



    finally:

        # closing database connection.

        if connection:
            cursor2.close()

            cursor1.close()

            connection.close()

            print("SQL connection is closed")

    return total_eng_users_cnt, total_new_users_cnt


def get_monthly_data():
    list_log_record = ''
    connection = ''
    """ query data from the log_record table """

    try:

        connection = get_connection()

        cursor = connection.cursor()

        cursor1 = connection.cursor()

        cursor2 = connection.cursor()

        select_Query = 'SELECT * from dbo.monthly_data_view;'
        cursor.execute(select_Query)

        log_record = cursor.fetchall()

        eng_select_Query = 'SELECT * from dbo.monthly_engaged_users_view;'
        cursor1.execute(eng_select_Query)

        engaged_u = cursor1.fetchall()

        new_select_Query = 'SELECT * from dbo.monthly_new_users_view;'
        cursor2.execute(new_select_Query)

        new_u = cursor2.fetchall()

        list_log_record = [list(tp) for tp in log_record]
        # print("into get_monthly_data = > ", list_log_record)


    except pyodbc.Error as error:

        print("Error while fetching data from SQL", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

    return list_log_record, engaged_u, new_u


def get_daily_data():
    engaged_u = 0

    new_u = 0
    connection = ''
    list_log_record = ''

    """ query data from the log_record table """

    try:

        connection = get_connection()

        cursor = connection.cursor()

        cursor1 = connection.cursor()

        cursor2 = connection.cursor()

        SQL_select_Query = 'SELECT * from dbo.daily_data_view;'
        cursor.execute(SQL_select_Query)
        log_record = cursor.fetchall()
        # print("in daily data ", log_record)

        eng_select_Query = 'SELECT * from dbo.engaged_users_view1;'
        cursor1.execute(eng_select_Query)

        engaged_u = cursor1.fetchall()

        new_select_Query = 'SELECT * from dbo.new_users_view1;'
        cursor2.execute(new_select_Query)
        new_u = cursor2.fetchall()
        # print("new_select_Query =>  ", new_u)

        list_log_record = [list(tp) for tp in log_record]

    except pyodbc.Error as error:

        print("Error while fetching data from SQL", error)



    finally:

        # closing database connection.

        if connection:
            cursor.close()

            connection.close()

            print("SQL connection is closed")

    # engaged_u, new_u =  get_total_users_cnt()
    return list_log_record, engaged_u, new_u


def get_total_event_cnt():
    global total_desc_cnt

    connection = ''

    try:

        connection = get_connection()

        cursor = connection.cursor()

        # postgreSQL_select_Query = 'SELECT * from chatbot.total_cnt_view_mview;'

        select_Query = 'SELECT * from dbo.total_ans_cnt_view;'

        cursor.execute(select_Query)

        # print("Selecting rows from chatbot_log table using cursor.fetchall")

        # log_record = cursor.fetchone()

        desc_total_cnt = cursor.fetchall()

        total_desc_cnt = [list(tp) for tp in desc_total_cnt]

    except pyodbc.Error as error:

        print("Error while fetching total_desc_cnt from SQL", error)

    finally:

        # closing database connection.

        if connection:
            cursor.close()

            connection.close()

            print("SQL connection is closed")

    return total_desc_cnt


@csrf_protect
def daily_charts(request):
    users_list = User.objects.filter(is_superuser=False)

    data, en, ne = get_daily_data()

    # print("ne and en = ", en, ne)
    tot_ans_cnt = get_total_event_cnt()
    # print("tot_ans_cnt_wr =============== ", tot_ans_cnt[0])

    busy_count = get_busy_period_count()

    busy_count_list = []

    for i in busy_count:

        if len(i[0]) > 5:
            busy_count_list.append(i)

    repeated_users = get_repeated_bot_users()

    reset, reset_month = reset_count()

    live, live_month = live_count()

    # print("""================================= in daily charts =================================""")

    event_categories = list()

    wrong_ans_data = list()

    right_ans_data = list()

    no_ans_data = list()

    for dt in data:

        if dt[1].__str__() not in event_categories:
            event_categories.append(dt[1].__str__())  # for answer desc

    edict = {}

    for i in event_categories:

        temp = [0, 0, 0]

        if i not in edict.keys():

            for dt in data:

                if i == dt[1].__str__():

                    if dt[0] == 'Right Answer': temp[0] = dt[2]

                    if dt[0] == 'Wrong Answer': temp[1] = dt[2]

                    if dt[0] == 'No Answer': temp[2] = dt[2]

        edict[i] = temp

    print(edict)

    for key, value in edict.items():
        right_ans_data.append(value[0])

        wrong_ans_data.append(value[1])

        no_ans_data.append(value[2])

    print("right_ans_data >>>> ", right_ans_data)
    print("wrong_ans_data >>>> ", wrong_ans_data)
    print("no_ans_data >>>> ", no_ans_data)

    wrong_answer = {

        'name': 'wrong_answer',

        'data': wrong_ans_data,

        'color': '#e53935'

    }

    right_answer = {

        'name': 'right_answer',

        'data': right_ans_data,

        'color': '#43a047'

    }

    no_answer = {

        'name': 'no_answer',

        'data': no_ans_data,

        'color': '#fb8c00'

    }

    bar_chart = {

        'chart': {'type': 'column',

                  },

        'title': {'text': 'Chatbot Log Summary(Bar Chart)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    line_chart = {

        'chart': {'type': 'line'},

        'title': {'text': 'Chatbot Log Summary(Line Charts)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif',

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    bar_dump = json.dumps(bar_chart)

    line_dump = json.dumps(line_chart)

    # for d in reset:

    #     print("---", d[1])

    # busy_count.remove(i)

    # print("busy_count_list", busy_count_list)
    # print('reset', reset)

    context = {

        'reset_count': len(reset) if len(reset) == 0 else reset[0][1],

        'live_count': len(live) if len(live) == 0 else live[0][1],

        'repeated_users': repeated_users,

        'busy_count': busy_count_list,

        'bar_chart': bar_dump,

        'line_chart': line_dump,

        'no_answer': "No Answer", 'no_answer_count': no_ans_data[0],

        'rt_answer': "Right Answer", 'rt_answer_count': right_ans_data[0],

        'wr_answer': "Wrong Answer", 'wr_answer_count': wrong_ans_data[0],

        'total_users': en[0][0] + len(ne),

        'engaged_users': en[0][0],

        'new_users': len(ne),

        'users_list': users_list

    }

    return context


def getUserById(_userid):
    users = User.objects.get(id=_userid)

    # print(users)


@csrf_protect
def monthly_charts(request):
    repeated_users = []

    reset, reset_month = [], []

    data, en, nu = get_monthly_data()

    tot_ans_cnt = get_total_event_cnt()

    _, live_month = live_count()

    repeated_users = get_repeated_bot_monthly_users()

    busy_period_count_month = get_busy_period_count_monthly()

    event_categories = list()

    wrong_ans_data = list()

    right_ans_data = list()

    no_ans_data = list()

    reset, reset_month = reset_count()

    # print("""================================= in monthly charts =================================""")
    for dt in data:

        if dt[1].__str__() not in event_categories:
            event_categories.append(dt[1].__str__())  # for answer desc

    edict = {}

    for i in event_categories:

        temp = [0, 0, 0]

        if i not in edict.keys():

            for dt in data:

                if i == dt[1].__str__():

                    if dt[0] == 'Right Answer': temp[0] = dt[2]

                    if dt[0] == 'Wrong Answer': temp[1] = dt[2]

                    if dt[0] == 'No Answer': temp[2] = dt[2]

        edict[i] = temp

    # print(edict)

    for key, value in edict.items():
        right_ans_data.append(value[0])

        wrong_ans_data.append(value[1])

        no_ans_data.append(value[2])

    # print("wr month", wrong_ans_data[0])

    # print("rt month", right_ans_data[0])

    # print("no month", no_ans_data[0])

    wrong_answer = {

        'name': 'wrong_answer',

        'data': wrong_ans_data,

        'color': '#e53935'

    }

    right_answer = {

        'name': 'right_answer',

        'data': right_ans_data,

        'color': '#43a047'

    }

    no_answer = {

        'name': 'no_answer',

        'data': no_ans_data,

        'color': '#fb8c00'

    }

    bar_chart = {

        'chart': {'type': 'column',

                  },

        'title': {'text': 'Chatbot Log Summary(Bar Chart)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Month',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    line_chart = {

        'chart': {'type': 'line'},

        'title': {'text': 'Chatbot Log Summary(Line Charts)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif',

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Month',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    bar_dump = json.dumps(bar_chart)

    line_dump = json.dumps(line_chart)

    temp_busy_count = []

    for i in busy_period_count_month:
        if i[0] != "": temp_busy_count.append(i)

    # print('[live_month info]', live_month)
    # print("bar_chart ", bar_dump)
    context = {

        'reset_month': len(reset_month) if len(reset_month) == 0 else reset_month[0][1],

        'repeated_users': repeated_users,

        'live_count': len(live_month) if len(live_month) == 0 else live_month[0][1],

        'busy_period_count_month': temp_busy_count,

        'total_users': len(en) + len(nu),

        'new_users': len(nu),

        'engaged_users': len(en),

        'bar_chart': bar_dump,

        'line_chart': line_dump,

        'no_answer': "No Answer", 'no_answer_count': no_ans_data[0],

        'rt_answer': "Right Answer", 'rt_answer_count': right_ans_data[0],

        'wr_answer': "Wrong Answer", 'wr_answer_count': wrong_ans_data[0],

    }

    return context


def create_excel(exl_name, col_desc, col_lst, request, rows):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Report.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(exl_name, cell_overwrite_ok=True)
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    font_style.font.bold = True
    cols_tab_desc = col_desc
    cols_head = col_lst
    font_style = xlwt.XFStyle()
    for col_num in range(len(cols_tab_desc)):
        ws.write(row_num, col_num, cols_tab_desc[col_num], font_style)

    for col_num in range(len(cols_head)):
        ws.write(row_num, col_num, cols_head[col_num], font_style)



    # print("================= # =========================", type(context['repeated_users'][0]))
    context = monthly_charts(request)
    rows = rows
    for row in rows:
        row_num += 1
        for row[0] in row:
            # print("================= # =========================",row[0])
            row_num += 1
            for col_num in range(len(row[0])):
                ws.write(row_num, col_num, str(row[0][col_num]), font_style)
                print("WOrked !!!")

@login_required
@user_passes_test(admin_check)
def report(request, report_name):
    # print('==============================', request.path)

    if request.path == '/export_pdf/Daily/':
        context = daily_charts(request)
        pdf = render_to_pdf('home/daily_export.html', context)
        if pdf:
            return HttpResponse(pdf, content_type='application/pdf')
        return HttpResponse("PDF Not Found.")
    elif request.path == '/export_pdf/Monthly/':
        context = monthly_charts(request)
        pdf = render_to_pdf('home/daily_export.html', context)
        if pdf:
            return HttpResponse(pdf, content_type='application/pdf')
        return HttpResponse("PDF Not Found.")

    elif request.path == '/export_excel/Daily/':
        # context = daily_charts(request)
        # print("context ", context)
        # response = HttpResponse(content_type='application/ms-excel')
        # response['Content-Disposition'] = 'attachment; filename="Report.xls"'
        # wb = xlwt.Workbook(encoding='utf-8')
        # ws = wb.add_sheet('Report')
        # row_num = 0
        # font_style = xlwt.XFStyle()
        # font_style.font.bold = True
        #
        # columns = ['Total Users', 'Engaged Users','New Users', 'Reset Count', 'Live Chat','No Answer',
        #            'Right Answer', 'Wrong Answer']

        # for col_num in range(len(columns)):
        #     ws.write(row_num, col_num, columns[col_num], font_style)
        #
        # font_style = xlwt.XFStyle()
        #
        # context = monthly_charts(request)
        #
        # rows = Log.objects.all().values_list(
        #     'event_type_id.description', 'user_email', 'event_question',
        #     'event_answer', 'user_datetime', 'intent')
        #
        # for row in rows:
        #     row_num += 1
        #     for col_num in range(len(row)):
        #         ws.write(row_num, col_num, str(row[col_num]), font_style)
        # wb.save(response)
        #
        # if wb:
        #     return response
        # return HttpResponse("No Data Found.")
        pass
    elif request.path == '/export_excel/Monthly/':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Report.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Report', cell_overwrite_ok=True)
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['Total Users', 'Engaged Users', 'New Users', 'Reset Count', 'Live Chat', 'No Answer',
                   'Right Answer', 'Wrong Answer']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        font_style = xlwt.XFStyle()

        context = monthly_charts(request)

        rows_stats = [(context['total_users'], context['engaged_users'], context['new_users'], context['reset_month'],
                       context['live_count'], context['no_answer_count'], context['rt_answer_count'],
                       context['wr_answer_count'])]

        for row in rows_stats:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), font_style)

        # font_style.font.bold = True
        cols_tab = ['Table for Repeated Interacted Users Data']
        cols_repeated_users = ['User Email', 'Bot Interaction', 'Date Time']
        #
        # for col_num in range(len(cols_tab)):
        #     ws.write(3, col_num, cols_tab[col_num], font_style)
        #
        # for col_num in range(len(cols_repeated_users)):
        #     ws.write(4, col_num, cols_repeated_users[col_num], font_style)
        #
        # font_style = xlwt.XFStyle()
        #
        # # print("================= # =========================", type(context['repeated_users'][0]))

        #
        # for row in rows_repeated_users:
        #     row_num += 3
        #     for row[0] in row:
        #         # print("================= # =========================",row[0])
        #         row_num += 1
        #         for col_num in range(len(row[0])):
        #             ws.write(row_num, col_num, str(row[0][col_num]), font_style)
        context = monthly_charts(request)
        rows_repeated_users = [(context['repeated_users'])]
        create_excel('repeated_users_data', cols_tab, cols_repeated_users, request, rows_repeated_users)

        # busiest period table
        font_style.font.bold = True
        cols_tab = ['Table for Busiest period of Chatbot interaction']
        cols_busy_period_count = ['Department', 'Count', 'Date Time']

        for col_num in range(len(cols_tab)):
            ws.write(3, col_num, cols_tab[col_num], font_style)

        for col_num in range(len(cols_busy_period_count)):
            ws.write(4, col_num, cols_busy_period_count[col_num], font_style)

        rows_busy_prd_cnt_month = [(context['busy_period_count_month'])]
        # print(context['busy_period_count_month'])
        # print("================= # =========================", type(context['busy_period_count_month'][0]))

        for row in rows_busy_prd_cnt_month:
            row_num += 3
            for row[0] in row:
                # print("================= # =========================",row[0])
                row_num += 1
                for col_num in range(len(row[0])):
                    ws.write(row_num, col_num, str(row[0][col_num]), font_style)

        wb.save(response)

        if wb:
            return response
        return HttpResponse("No Data Found.")

        # context = monthly_charts(request)
        # rows = [context['total_users'], context['engaged_users'], context['new_users'], context['reset_month'],
        #         context['live_count'], context['no_answer_count'], context['rt_answer_count'],
        #         context['wr_answer_count']]
        #
        # columns = ['Total Users', 'Engaged Users', 'New Users', 'Reset Count', 'Live Chat', 'No Answer',
        #            'Right Answer', 'Wrong Answer']
        #
        # df = pd.DataFrame(rows, columns=columns)
        # df.to_excel('./states.xlsx', sheet_name='States', index=False)

    else:

        if report_name == "Daily":

            context = daily_charts(request)

            return render(request, 'home/user_report.html', context)

        elif report_name == "Monthly":
            # print("into elif before monthly charts", report_name)

            context = monthly_charts(request)
            # print("in reports context - > ", context)

            return render(request, 'home/month_report.html', context)


@login_required
@user_passes_test(admin_check)
def users(request):
    reports = Report.objects.values("assigned_to", "report_name")
    users_list = User.objects.filter(is_superuser=False)

    dept_usr_list = DepartmentAdminUser.objects.values("user", "usertype", "department")
    usertype_lst = DepartmentAdminUser.objects.all()  # added
    department = Department.objects.all()  # added

    ulist = []
    for u in usertype_lst:
        if u.user.username not in ulist:
            ulist.append(u.user.username)
            ulist.append(u.usertype.usertype)

    temp_list = []
    temp_dp_list = []
    for dpt in department:
        # print("dpt=> ", dpt.department)
        if dpt.department not in temp_dp_list:
            temp_dp_list.append(dpt.id)
            temp_dp_list.append(dpt.department)
            temp_list.append(dpt.department)

    # for d in dept_usr_list:
    #     print("-?>", d.department)

    _user_reports = {}
    for user in users_list:
        _temp = []
        for dptl in dept_usr_list:
            # print("Dictionary-", dptl)
            if user.id == dptl['user']:
                # print("################################")
                # print("", user.id, dptl['user'])
                # print("", dptl['department'])
                # print(type(dptl['department']))
                # print("", temp_dp_list.index(dptl['department']))
                # print("", temp_dp_list.index(dptl['department']) + 1)

                _temp.append(temp_dp_list[temp_dp_list.index(dptl['department']) + 1])

        if user.id not in _user_reports:
            _user_reports[user.id] = _temp
    # print("user_reports - ", _user_reports)
    if len(users_list) == 0:
        html = ""
        html += '<tr><td colspan=3>No Users Found</td></tr>'
        context = {
            'report_html': html,
        }
        return render(request, 'home/user_admin.html', context)
    else:
        html = ""
        for user in range(len(users_list)):
            ulist_var = ulist[ulist.index(users_list[user].username) + 1]
            html += '<tr><td>' + str(user + 1) + '</td><td>' + users_list[
                user].username + '<td>' + str(ulist_var) + '</td>' + '</td><td><button class="dropbtn" id="' + str(
                user + 1) + '" onclick="show(' + str(
                user + 1) + ')" value="Reports Assigned">Departments Assigned</button></td>'
            for tmp in temp_list:
                html += '<tr style="text-align:center" class="show' + str(
                    user + 1) + '" id="reports_hide"><td  colspan="3">' + tmp + '</td>'
                inserted = False
                for key, value in _user_reports.items():
                    # print("key value ",  key, value)
                    if users_list[user].id == key:
                        for i in value:
                            if i == tmp:
                                inserted = True
                                html += '<td><input type="checkbox" style="margin-right: 180px;" value="assigned" onclick=check(event,' + str(
                                    users_list[user].id) + ',"' + tmp + '") checked/></td>'
                if not inserted:
                    html += '<td><input type="checkbox" style="margin-right: 180px;" value="assigned" onclick=check(event,' + str(
                        users_list[user].id) + ',"' + tmp + '") /></td>'
                html += '</tr>'
            html += '</tr>'
        context = {
            'reports': temp_list,
            'user_list': users_list,
            'report_html': html,
        }
        return render(request, 'home/user_admin.html', context)


def update_report(request):
    # print('############## update_report #################')
    # print("1.", request.POST)
    # print("2.", request.POST['userid'])
    # print("3.", request.POST['reportname'])
    # print("4.", request.POST['reportname'][0])
    userid_ = User.objects.get(id=request.POST['userid'])
    userid_name = userid_.username
    # print("userid_ ", userid_)
    # print("userid_username  ", userid_name)

    ulist = []
    usertype_lst = DepartmentAdminUser.objects.all()
    for u in usertype_lst:
        if u.user.username not in ulist:
            ulist.append(u.user.username)
            ulist.append(u.usertype.usertype)

    u_type = ulist[ulist.index(userid_name) + 1]
    # print("ulist - ", ulist)
    # print("utype - ", u_type)
    # print("utype type - ", type(u_type))

    report_nm = request.POST["reportname"]
    # print("report_nm - ", report_nm)
    # print("type - ", type(report_nm))
    did = Department.objects.get(department=report_nm)
    uid = UserType.objects.get(usertype=u_type)
    # print("did ================> ", did)
    # print("uid ================> ", uid)
    # userid_, uid, did,

    if request.POST['checked'] == 'true':
        DepartmentAdminUser.objects.create(user=userid_, usertype=uid, department=did)
    else:
        instance = DepartmentAdminUser.objects.get(user=userid_, usertype=uid, department=did)
        print(">>> ", instance)
        instance.delete()
    return redirect('user/')


@csrf_protect
@login_required
@user_passes_test(admin_check)
def index(request):
    current_user = request.user
    # print(current_user.id)
    # print(current_user.username)
    users_list = User.objects.filter(is_superuser=False)

    if current_user.username == "admin":

        reports = Report.objects.values("assigned_to", "report_name")
        departments = Department.objects.all()

        temp_list = []

        for i in reports:

            if i["report_name"] not in temp_list:
                temp_list.append(i["report_name"])

        return render(request, 'home/index_admin.html', {'userlist': users_list, 'reports': temp_list,
                                                         'depts': departments})

    else:

        reports = Report.objects.values("assigned_to", "report_name")

        temp_list = []

        for i in reports:

            if i["report_name"] not in temp_list and i['assigned_to'] == current_user.id:
                temp_list.append(i["report_name"])

        return render(request, 'home/user_adoreta.html', {'reports': temp_list})


@csrf_protect
@login_required(login_url="/login/")
def pages(request):
    context = {}

    try:

        load_template = request.path.split('/')[-1]

        # print('load_template', load_template)

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))

        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)

        return HttpResponse(html_template.render(context, request))



    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')

        return HttpResponse(html_template.render(context, request))

    except:

        html_template = loader.get_template('home/page-500.html')

        return HttpResponse(html_template.render(context, request))


@csrf_protect
@login_required(login_url="/admin/")
def index_admin(request):
    context = {'segment': 'index_admin'}

    html_template = loader.get_template('home/index_admin.html')

    return HttpResponse(html_template.render(context, request))


@csrf_protect
@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dept_register(request):
    session_dept = request.session['depart']
    reg_user = request.user
    if reg_user is not None:
        name_title = reg_user.username
        chk_super = User.objects.filter(username=name_title, is_superuser=True).exists()
        if chk_super:
            depart_name = "Super_Admin"
            admin_type = 'admin'
        else:
            did = Department.objects.get(department=session_dept)
            user_ = DepartmentAdminUser.objects.get(user=reg_user, department=did)
            admin_type = user_.usertype.usertype
            depart_name = user_.department.department

            # print("admin_type => ", admin_type)
            # print("depart_name ", depart_name)
            # print("type -admin ", type(admin_type))

            if admin_type == 'Department_User':
                # print("=inside if ", admin_type)
                return redirect('page_not_found')

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm = request.POST.get('confirm')
        utype = request.POST.get('utype')
        dtype = request.POST.getlist('dtype')
        # print("dtype-> ", dtype)
        if len(dtype) == 0:
            dtype = [session_dept]

        list_usertype = UserType.objects.all()
        list_dept = Department.objects.all()
        # list_email = Department.objects.all().values('user_email')
        if password != confirm:
            color = "red"
            context = {
                'color': color,
                'depart_name': depart_name,
                'name_title': name_title,
                'admin_type': admin_type,
                'msg': 'PASSWORD & CONFIRM PASSWORD DID NOT MATCH !',
                'list_usertype': list_usertype,
                'list_dept': list_dept,
            }
            return render(request, 'accounts/register_.html', context)
        else:
            # print("department=> ", dtype)

            # create django username
            user = User.objects.create(username=username, email=email)
            user.set_password(password)
            user.save()

            for dt in dtype:
                did = Department.objects.get(department=dt)
                uid = UserType.objects.get(usertype=utype)
                dau = DepartmentAdminUser.objects.create(user=user, usertype=uid, department=did)

                # print("dt========", dt, did, username, email, password, utype)
                dau.save()

            # print("User Created!!!")
            color = 'green'
            context = {
                'color': color,
                'depart_name': depart_name,
                'name_title': name_title,
                'admin_type': admin_type,
                'msg': f'ACCOUNT CREATED FOR {username} !',
                'list_usertype': list_usertype,
                'list_dept': list_dept,
            }
            return render(request, 'accounts/register_.html', context)

    list_usertype = UserType.objects.all()
    list_dept = Department.objects.all()

    context = {
        'depart_name': depart_name,
        'name_title': name_title,
        'admin_type': admin_type,
        'list_usertype': list_usertype,
        'list_dept': list_dept,
    }

    return render(request, 'accounts/register_.html', context)


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def get_dept_data(request, department_name):
    dept_list = Department.objects.all()
    updated_list = []

    # print("department==? ", department_name)

    for dl in dept_list:
        updated_list.append(dl.department)

    if department_name not in updated_list:
        return redirect('page_not_found')
    else:
        depart_user = request.user
        if depart_user.is_superuser:
            name_title = depart_user.username
            depart_name = "Super_Admin"
            admin_type = "admin"
        else:
            did = Department.objects.get(department=department_name)
            user_ = DepartmentAdminUser.objects.filter(user=depart_user, department=did)
            # print("user_ = > ", user_)
            if len(user_) == 0:
                return redirect('page_not_found')

            name_title = depart_user.username
            depart_name = user_[0].department.department
            admin_type = user_[0].usertype.usertype

        wr_answer_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='3').count()
        rt_answer_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='4').count()
        no_answer_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='5').count()
        live_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='6').count()
        reset_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='7').count()
        total_users_cnt = Log.objects.all().values('user_email').distinct().filter(intent=department_name).count()
        eng_users_cnt = Log.objects.all().values('user_email').filter(intent=department_name).count()
        repeated_users_list = Log.objects.all().values('user_email', 'user_datetime', 'intent')

        new_user_data = Log.objects.all().values('user_email').distinct().filter(intent=department_name)
        # new_user_data = Log.objects.all().values('user_email').filter(intent=department_name).count()

        print("=> log_dept ", len(new_user_data), eng_users_cnt)

        # print("dept counts_data=> ",
        #       "wr_answer_cnt-", wr_answer_cnt,
        #       "rt_answer_cnt-", rt_answer_cnt,
        #       "no_answer_cnt-", no_answer_cnt,
        #       "live_cnt-", live_cnt,
        #       "reset_cnt-", reset_cnt,
        #       "total_users_cnt-", total_users_cnt,
        #       "eng_users_cnt-", eng_users_cnt)

        ## bar and line chart
        bar_data, line_data, rep_data, busy_data, new_user_data__ = dept_chart(department_name)
        print("eng_users_cnt= ", eng_users_cnt)

        context = {
            'admin_type': admin_type,
            'depart_name': depart_name,
            'name_title': name_title,
            'dept_name': department_name,
            'wr_answer': 'Wrong Answer', 'wr_answer_count': wr_answer_cnt,
            'rt_answer': 'Right Answer', 'rt_answer_count': rt_answer_cnt,
            'no_answer': 'No Answer', 'no_answer_count': no_answer_cnt,
            'live_count': live_cnt,
            'reset_count': reset_cnt,
            'total_users': total_users_cnt,
            'eng_users_cnt': eng_users_cnt,
            'repeated_users_list': repeated_users_list,
            'bar_chart_data': bar_data,
            'line_chart_data': line_data,
            'repeated_users_data': rep_data,
            'busy_users_data': busy_data,
            'new_user_data': len(new_user_data),
        }

    return render(request, 'home/department.html', context)


@csrf_protect
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dept_login(request):
    global admin_type, depart_name
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        dtype = request.POST.get('dtype')
        # print("dtype-> ", dtype)
        # print("==", type(dtype))

        user = authenticate(username=username, password=password)
        if user is not None:
            chk_super = User.objects.filter(username=username, is_superuser=True).exists()
            did = Department.objects.get(department=dtype)
            check_dpu = DepartmentAdminUser.objects.filter(user=user, department=did).exists()
            # print("check_dpu - ", check_dpu)
            if chk_super:
                login(request, user)
                request.session['depart'] = dtype
                return redirect('department_data', department_name=dtype)
            elif check_dpu:
                login(request, user)
                request.session['depart'] = dtype
                return redirect('department_data', department_name=dtype)
            else:
                msg = f'YOU ARE NOT REGISTERED TO {dtype} DEPARTMENT'
                list_dept = Department.objects.all()
                context = {
                    'msg': msg,
                    'list_dept': list_dept
                }
                return render(request, 'accounts/login_.html', context)
        else:
            msg = 'INVALID CREDENTIALS'
            list_dept = Department.objects.all()
            context = {
                'msg': msg,
                'list_dept': list_dept
            }
            return render(request, 'accounts/login_.html', context)

    list_dept = Department.objects.all()
    context = {
        'list_dept': list_dept
    }
    return render(request, 'accounts/login_.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dept_logout(request):
    logout(request)
    return redirect('login_')
